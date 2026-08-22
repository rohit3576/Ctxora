"""Binary-format parsers: PDF, DOCX, XLSX.

These wrap libraries without complete type stubs (pymupdf, python-docx,
openpyxl). The rest of the package sees strictly typed ChunkedPage values;
the untyped-library handling is accepted inside this module.
"""

import io

from rag.chunker import ChunkedPage


def pdf_pages(content: bytes) -> tuple[ChunkedPage, ...]:
    """Page-per-page text extraction via pymupdf."""
    import pymupdf

    pages: list[ChunkedPage] = []
    with pymupdf.open(stream=content, filetype="pdf") as document:
        for number, page in enumerate(document, start=1):
            text = page.get_text("text").strip()
            if text:
                pages.append(ChunkedPage(page_number=number, sections=(("Page", text),)))
    return tuple(pages)


def docx_sections(content: bytes) -> tuple[ChunkedPage, ...]:
    """Paragraph extraction via python-docx; heading styles become sections."""
    from docx import Document

    document = Document(io.BytesIO(content))
    sections: list[tuple[str, str]] = []
    current_title = "Document"
    buffer: list[str] = []
    for paragraph in document.paragraphs:
        piece = paragraph.text.strip()
        if not piece:
            continue
        style = paragraph.style
        style_name = style.name if style is not None and style.name else ""
        if style_name.startswith("Heading"):
            if buffer:
                sections.append((current_title, " ".join(buffer)))
                buffer = []
            current_title = piece
        else:
            buffer.append(piece)
    if buffer:
        sections.append((current_title, " ".join(buffer)))
    return (ChunkedPage(page_number=1, sections=tuple(sections)),)


def xlsx_pages(content: bytes) -> tuple[ChunkedPage, ...]:
    """Sheet-per-page extraction via openpyxl."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    pages: list[ChunkedPage] = []
    for number, sheet in enumerate(workbook.worksheets, start=1):
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
        if lines:
            title = sheet.title or "Sheet"
            pages.append(ChunkedPage(page_number=number, sections=((title, "\n".join(lines)),)))
    return tuple(pages)
